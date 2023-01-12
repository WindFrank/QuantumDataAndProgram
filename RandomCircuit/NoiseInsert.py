import sys

import cirq
from ProgramSet.examples.basic_arithmetic import circuit_noise_add, circuit_add_other_noise, write_to_excel
from RandomCircuit import read_circuit_code
from openpyxl import load_workbook, Workbook
import numpy as np


def noise_insert(circuit_code, noise_type, p_or_gamma, count):
    i = 0
    success_times = 0
    c = read_circuit_code(circuit_code)
    sim = cirq.Simulator()
    while i < count:
        result = sim.run(c)
        c_noise = circuit_noise_add(c, noise_type, p_or_gamma)
        result_noise = sim.run(c_noise)
        if result == result_noise:
            success_times += 1
        i += 1
    return success_times / count


def noise_sim(circuit_code, noise_type, p_or_gamma):
    i = 0
    success_times = 0
    c = read_circuit_code(circuit_code)
    sim = cirq.Simulator()
    result = sim.run(c)
    result_noise = circuit_add_other_noise(c, noise_type, p_or_gamma, 1)
    if result == result_noise:
        success_times += 1
    i += 1
    return success_times


def main():
    data = []
    circuit_codes = read_xls('satisfy_circuit.xlsx')
    circuit_codes.pop(0)
    count = 0
    for item in circuit_codes:
        circuit_code = item[0]
        repetitions = 1000
        p_or_gammas = np.linspace(0, 0.5, 51, endpoint=True)
        row_data = []
        for noise_type in ['b', 'd', 'pf']:
            for p_or_gamma in p_or_gammas:
                success_rate = noise_insert(circuit_code, noise_type, p_or_gamma, repetitions)
                row_data.append(success_rate)
                sys.stdout.write(f'circuit_code:{circuit_code} noise_type:{noise_type} p/gamma:{p_or_gamma} success_rate:{success_rate}')
                sys.stdout.flush()
                print()
        data.append(row_data)
        count += 1
        write_to_excel(f'circuit_noise_success_rate{count}.xlsx', 'sheet1', 'info', data)


def read_xls(location):
    excel = load_workbook(location)
    # 获取sheet：
    table = excel.get_sheet_by_name('sheet1')  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    cols = table.max_column  # 获取列数
    # 获取单元格值：
    data = []
    for i in range(rows):
        row_data = []
        for j in range(cols):
            row_data.append(table.cell(row=i + 1, column=j + 1).value)  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
        data.append(row_data)
    return data



if __name__ == '__main__':
    main()

