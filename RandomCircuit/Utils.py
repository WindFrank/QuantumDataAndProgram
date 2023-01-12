import openpyxl
from openpyxl import load_workbook

from xlsTest import write_to_excel


def clear_repeat(file_path):
    circuits_info = read_xls(file_path)
    last_info = ''
    remove_list = []
    for circuit_info, i in zip(circuits_info, range(len(circuits_info))):
        if i == 0:
            continue
        elif len(circuit_info[0]) <= 4:
            remove_list.append(i)
            continue
        elif last_info == '':
            last_info = circuit_info[0]
            continue
        temp = circuit_info[0]
        if temp == last_info or temp[len(temp) - 2] != '6':
            remove_list.append(i)
        last_info = circuit_info[0]
    remove_list.reverse()
    for i in remove_list:
        circuits_info.pop(i)
    print(f'processed data:{len(circuits_info) - 1}')
    write_to_excel(f'unrepeated_zkd.xlsx', 'sheet1', 'info', circuits_info)


def clear_repeat_zkd(file_path):
    circuits_info = read_xls(file_path)
    last_info = ''
    remove_list = []
    for circuit_info, i in zip(circuits_info, range(len(circuits_info))):
        if i == 0:
            continue
        elif last_info == '':
            last_info = circuit_info[0]
            continue
        temp = circuit_info[0]
        if temp == last_info:
            remove_list.append(i)
        last_info = circuit_info[0]
    remove_list.reverse()
    for i in remove_list:
        circuits_info.pop(i)
    print(f'processed data:{len(circuits_info) - 1}')
    write_to_excel(f'GNN_training_data_new_gate_e1.xlsx', 'sheet1', 'info', circuits_info)


def read_xls(location, sheet='sheet1'):
    excel = load_workbook(location)
    # 获取sheet：
    table = excel.get_sheet_by_name(sheet)  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    cols = table.max_column  # 获取列数
    # 获取单元格值：
    data = []
    for i in range(rows):
        row_data = []
        for j in range(cols):
            row_data.append(table.cell(row=i + 1, column=j + 1).value)  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
        data.append(row_data)
    return data


def read_str_vector(str_vector):
    str_vector = str_vector.replace('[', '')
    str_vector = str_vector.replace(']', '')
    str_vector = str_vector.replace(' ', '')
    result = []
    i = 0
    while i < len(str_vector):
        [number, i] = find_double_number(str_vector, i)
        result.append(number)
        i = i + 1
    return result


def find_double_number(circuit_string, index):
    i = index
    temp = 0
    integer_temp = 0
    tag = 1
    coefficient = 1
    float_flag = False
    if circuit_string[i] == '-':
        tag = -1
        i += 1
    while True:
        if i < len(circuit_string) and '0' <= circuit_string[i] <= '9':
            temp = temp * 10
            temp += int(circuit_string[i])
            if float_flag:
                coefficient *= 0.1
        elif i < len(circuit_string) and circuit_string[i] == '.':
            float_flag = True
            integer_temp = temp
            temp = 0
        else:
            break
        i = i + 1
    return [(integer_temp + temp * coefficient) * tag, i]


if __name__ == '__main__':
    clear_repeat_zkd('GNN_training_data_18000_new_gate_simulate_ALL.xlsx')
