import sys
from typing import Tuple, Iterable

import cirq
import random
import torch
from RandomCircuit.RandomNoise import RandomNoise
from openpyxl import load_workbook

from xlsTest import write_to_excel
from torch_geometric.data import Data


def random_circuit_generate_no_gate_number(pre_width, pre_depth):
    width = pre_width
    depth = pre_depth
    qubits = [cirq.GridQubit(i, 0) for i in range(width)]
    temp_gate_number = 0
    temp_multi_gate_number = 0
    depths = [0 for i in range(width)]
    judgement = 0
    gate_type = []
    circuit = cirq.Circuit()
    string_code = [width]  # 先放门位数，再放门类型，再放作用量子位，若有多个则依次放入，无顺序
    while judgement == 0:
        if max(depths) >= depth:
            judge = random.random()
            if judge > 0.5:
                break
        gate_type = [1 for i in range(3)]
        gate_type.append(2)
            #gate_type.append(3)
        random_type = gate_type[random.randint(0, len(gate_type) - 1)]
        gates = []
        if random_type == 1:
            qubit_choose = random.randint(0, width - 1)
            qubit = cirq.GridQubit(qubit_choose, 0)
            gates.append(cirq.X(qubit))
            gates.append(cirq.Y(qubit))
            gates.append(cirq.Z(qubit))
            gates.append(cirq.H(qubit))
            '''gates.append(cirq.S(qubit))
            gates.append(cirq.T(qubit))'''
            gate_choose = random.randint(0, len(gates) - 1)
            final_gate = gates[gate_choose]
            circuit.append(final_gate)
            depths[qubit_choose] += 1
            string_code.append(1)
            string_code.append(gate_choose)
            string_code.append(qubit_choose)
        elif random_type == 2 and width >= 2:
            qubit_choose1 = random.randint(0, width - 1)
            qubit_choose2 = -1
            while qubit_choose2 <= -1 or qubit_choose2 >= width:
                qubit_choose2 = qubit_choose1 + 1
                if random.random() > 0.5:
                    qubit_choose2 = qubit_choose1 - 1

            # qubit_choose2 = random.randint(0, width - 1)
            # while qubit_choose1 == qubit_choose2:
            #     qubit_choose2 = random.randint(0, width - 1)
            qubit1 = cirq.GridQubit(qubit_choose1, 0)
            qubit2 = cirq.GridQubit(qubit_choose2, 0)
            together_depth = max(depths[qubit_choose1], depths[qubit_choose2])
            if together_depth == depth:
                continue
            gates.append(cirq.CZ(qubit1, qubit2))
            # gate_choose = random.randint(0, len(gates) - 1)
            final_gate = gates[0]
            depths[qubit_choose1] = together_depth + 1
            depths[qubit_choose2] = together_depth + 1
            circuit.append(final_gate)
            temp_multi_gate_number += 1
            string_code.append(2)
            string_code.append(0)
            string_code.append(qubit_choose1)
            string_code.append(qubit_choose2)
        temp_gate_number += 1
    circuit_code = ''
    for i in range(len(string_code)):
        circuit_code += str(string_code[i])
    string_code = circuit_code

    circuit = add_measurement(circuit, qubits)
    sim = cirq.Simulator()
    result = sim.run(circuit, repetitions=1000)
    last_list = []
    same = [1 for i in range(width)]
    key_string = ''
    for i in range(width):
        if i == 0:
            key_string += f'({i}, 0)'
        else:
            key_string += f',({i}, 0)'
    for my_list in result.measurements[key_string]:
        if len(last_list) == 0:
            last_list = my_list
            continue
        bool_judge = last_list == my_list
        for is_same, i in zip(bool_judge, range(width)):
            if is_same == 0:
                same[i] = 0
        if any(same) == 0:
            break

    measure_count = 0
    if any(same) == 1:
        for bool_judge, index in zip(same, range(width)):
            if bool_judge == 1:
                string_code = string_code + str(1) + str(6) + str(index)
                measure_count += 1
        string_code, width, depth, temp_gate_number, temp_multi_gate_number, measure_count = wash_circuit(string_code)
        if string_code == '':
            return random_circuit_generate_no_gate_number(pre_width, pre_depth)
    test_cut_line = circuit_cut(string_code)
    last_gate = test_cut_line[len(test_cut_line) - 1]
    if len(last_gate) != 3 or last_gate[len(last_gate) - 2] != 6 or test_cut_line[1][1] == 6:
        return random_circuit_generate_no_gate_number(pre_width, pre_depth)
    return string_code, width, depth, temp_gate_number, temp_multi_gate_number, measure_count


def random_circuit_generate(width, depth, gate_number, multi_gate_number):
    if gate_number < multi_gate_number:
        raise Exception('Gate number less than multi gate number.')
    if width * depth < gate_number:
        raise Exception('Width and depth is so small that cannot contain such number of gates.')
    qubits = [cirq.GridQubit(i, 0) for i in range(width)]
    temp_gate_number = 0
    temp_multi_gate_number = 0
    depths = [0 for i in range(width)]
    judgement = 0
    gate_type = []
    circuit = cirq.Circuit()
    string_code = [width]  # 先放门位数，再放门类型，再放作用量子位，若有多个则依次放入，无顺序
    while judgement == 0:
        if max(depths) == depth and temp_gate_number == gate_number and temp_multi_gate_number == multi_gate_number:
            break
        elif depths == [depth for i in range(width)] or temp_gate_number == gate_number:
            return random_circuit_generate(width, depth, gate_number, multi_gate_number)
        if temp_multi_gate_number == multi_gate_number:
            gate_type = [1]
        else:
            gate_type = [1 for i in range(int(2 * gate_number / multi_gate_number))]
            gate_type.append(2)
            #gate_type.append(3)
        random_type = gate_type[random.randint(0, len(gate_type) - 1)]
        gates = []
        if random_type == 1:
            qubit_choose = random.randint(0, width - 1)
            qubit = cirq.GridQubit(qubit_choose, 0)
            gates.append(cirq.X(qubit))
            gates.append(cirq.Y(qubit))
            gates.append(cirq.Z(qubit))
            gates.append(cirq.H(qubit))
            '''gates.append(cirq.S(qubit))
            gates.append(cirq.T(qubit))'''
            gate_choose = random.randint(0, len(gates) - 1)
            final_gate = gates[gate_choose]
            circuit.append(final_gate)
            depths[qubit_choose] += 1
            string_code.append(1)
            string_code.append(gate_choose)
            string_code.append(qubit_choose)
        elif random_type == 2 and width >= 2:
            qubit_choose1 = random.randint(0, width - 1)
            qubit_choose2 = -1
            while qubit_choose2 <= -1 or qubit_choose2 >= width:
                qubit_choose2 = qubit_choose1 + 1
                if random.random() > 0.5:
                    qubit_choose2 = qubit_choose1 - 1

            # qubit_choose2 = random.randint(0, width - 1)
            # while qubit_choose1 == qubit_choose2:
            #     qubit_choose2 = random.randint(0, width - 1)
            qubit1 = cirq.GridQubit(qubit_choose1, 0)
            qubit2 = cirq.GridQubit(qubit_choose2, 0)
            together_depth = max(depths[qubit_choose1], depths[qubit_choose2])
            if together_depth == depth:
                continue
            gates.append(cirq.CZ(qubit1, qubit2))
            #gate_choose = random.randint(0, len(gates) - 1)
            final_gate = gates[0]
            depths[qubit_choose1] = together_depth + 1
            depths[qubit_choose2] = together_depth + 1
            circuit.append(final_gate)
            temp_multi_gate_number += 1
            string_code.append(2)
            string_code.append(0)
            string_code.append(qubit_choose1)
            string_code.append(qubit_choose2)
        elif random_type == 3 and width >= 3:
            qubit_choose1 = random.randint(0, width - 1)
            qubit_choose2 = random.randint(0, width - 1)
            while qubit_choose1 == qubit_choose2:
                qubit_choose2 = random.randint(0, width - 1)
            qubit_choose3 = random.randint(0, width - 1)
            while qubit_choose3 == qubit_choose2 or qubit_choose3 == qubit_choose1:
                qubit_choose3 = random.randint(0, width - 1)
            together_depth = max(depths[qubit_choose1], depths[qubit_choose2], depths[qubit_choose3])
            if together_depth == depth:
                continue
            qubit1 = cirq.GridQubit(qubit_choose1, 0)
            qubit2 = cirq.GridQubit(qubit_choose2, 0)
            qubit3 = cirq.GridQubit(qubit_choose3, 0)
            gates.append(cirq.TOFFOLI(qubit1, qubit2, qubit3))
            gates.append(cirq.CCZ(qubit1, qubit2, qubit3))
            gate_choose = random.randint(0, len(gates) - 1)
            final_gate = gates[gate_choose]
            circuit.append(final_gate)
            depths[qubit_choose1] = together_depth + 1
            depths[qubit_choose2] = together_depth + 1
            depths[qubit_choose3] = together_depth + 1
            temp_multi_gate_number += 1
            string_code.append(3)
            string_code.append(gate_choose)
            string_code.append(qubit_choose1)
            string_code.append(qubit_choose2)
            string_code.append(qubit_choose3)
        temp_gate_number += 1
    circuit_code = ''
    for i in range(len(string_code)):
        circuit_code += str(string_code[i])
    string_code = circuit_code

    circuit = add_measurement(circuit, qubits)
    sim = cirq.Simulator()
    result = sim.run(circuit, repetitions=1000)
    last_list = []
    same = [1 for i in range(width)]
    key_string = ''
    for i in range(width):
        if i == 0:
            key_string += f'({i}, 0)'
        else:
            key_string += f',({i}, 0)'
    for my_list in result.measurements[key_string]:
        if len(last_list) == 0:
            last_list = my_list
            continue
        bool_judge = last_list == my_list
        for is_same, i in zip(bool_judge, range(width)):
            if is_same == 0:
                same[i] = 0
        if any(same) == 0:
            break

    measure_count = 0
    if any(same) == 1:
        for bool_judge, index in zip(same, range(width)):
            if bool_judge == 1:
                string_code = string_code + str(1) + str(6) + str(index)
                measure_count += 1
        string_code, width, depth, gate_number, multi_gate_number = wash_circuit(string_code)
    return string_code, width, depth, gate_number, multi_gate_number, measure_count


def add_measurement(c, qubits):
    c.append(cirq.measure(*qubits))
    return c


def to_string(number_list):
    new_str = []
    for i in range(len(number_list)):
        new_str.append(str(number_list[i]))
    return ''.join(new_str)


def read_circuit_code(circuit_code, random_noise=0, channel=None):
    c = cirq.Circuit()
    codes = list(map(int, circuit_code))
    qubits = []
    i = 0
    while i < len(codes):
        if i == 0:
            for qubit_index in range(codes[i]):
                qubits.append(cirq.GridQubit(qubit_index, 0))
        else:
            gate_multi_number = codes[i]  # 读门位数类型
            if gate_multi_number == 1:
                i = i + 1
                gate_index = codes[i]  # 读门类型

                i = i + 1
                qubit = qubits[codes[i]]  # 读量子位
                gates = [cirq.X(qubit), cirq.Y(qubit), cirq.Z(qubit), cirq.H(qubit), cirq.S(qubit), cirq.T(qubit),
                         cirq.measure(qubit)]
                c.append(gates[gate_index])
                if gate_index == 6:
                    break
                if random_noise == 1 and gate_index != 6:
                    c.append(channel.on(qubit))
            elif gate_multi_number == 2:
                i = i + 1
                gate_index = codes[i]  # 读门类型
                i = i + 1
                qubit1 = qubits[codes[i]]  # 读量子位
                i = i + 1
                qubit2 = qubits[codes[i]]  # 读量子位
                gates = [cirq.CNOT(qubit1, qubit2), cirq.CZ(qubit1, qubit2)]
                c.append(gates[gate_index])
                if random_noise == 1:
                    c.append(channel.on(qubit1))
                    c.append(channel.on(qubit2))
            else:
                raise Exception('The circuit code is error.')
        i = i + 1
    return c


def node_get(circuit_code):
    cut_gates = circuit_cut(circuit_code)
    count = 0
    for gate in cut_gates:
        if len(gate) == 1:
            continue
        count += int(gate[0])
    return count


def trans_circuit_to_graph(circuit_code, right_rate):  # 将电路转换为图数据结构
    codes = list(map(int, circuit_code))
    node_count = node_get(circuit_code)
    i = 0
    x = []
    '''y = 1 if right_rate >= 0.5 else 0'''

    '''if right_rate >= 0.5:
        y[1] = 1
    else:
        y[0] = 1'''
    node_int_all = []
    edge_index = []
    depth_note = []
    pre_nodes = []
    node_index = 1
    while i < len(codes):
        if i == 0:
            for j in range(codes[0]):
                depth_note.append(0)
                pre_nodes.append(-1)
            i += 1
            x.append([0 for _ in range(38)])
            node_int_all.append([])
            continue
        nodes = []
        temp_node_int = []
        gate_qubit_number = codes[i]
        i += 1

        if gate_qubit_number == 1:
            gate_type = codes[i] + 1
            if gate_type == 7:
                gate_type = 5
            qubit = codes[i+1]
            depth_note[qubit] += 1
            pre_node = pre_nodes[qubit]
            pre_distance = 0
            pre_node_gate_type = -1
            if pre_node != -1:
                edge_index.append([pre_node, node_index])
                pre_distance = 1
                pre_node_gate_type = node_int_all[pre_node][0]
            else:
                edge_index.append([0, node_index])
            node = get_graph_node(gate_type, qubit, depth_note[qubit] - 1, pre_distance=pre_distance, pre_node=pre_node_gate_type + 1)
            node_int = [gate_type, qubit, depth_note[qubit] - 1]
            temp_node_int.append(node_int)
            nodes.append(node)
            pre_nodes[qubit] = node_index
            node_index += 1
            i += 2
        elif gate_qubit_number == 2:
            gate_type = codes[i]
            qubit1 = codes[i+1]
            qubit2 = codes[i+2]
            node1 = [-1, qubit1]
            node2 = [0 if gate_type == 0 else -1, qubit2]
            depth_max = max(depth_note[qubit1], depth_note[qubit2]) + 1
            node1.append(depth_max)
            node2.append(depth_max)
            pre_distance_1 = 0
            pre_distance_2 = 0
            pre_node_gate_type1 = -1
            pre_node_gate_type2 = -1
            if pre_nodes[qubit1] != -1:
                edge_index.append([pre_nodes[qubit1], node_index])
                pre_distance_1 = depth_max - depth_note[qubit1]
                pre_node_gate_type1 = node_int_all[pre_nodes[qubit1]][0]
            else:
                edge_index.append([0, node_index])
            pre_nodes[qubit1] = node_index
            if pre_nodes[qubit2] != -1:
                edge_index.append([pre_nodes[qubit2], node_index + 1])
                pre_distance_2 = depth_max - depth_note[qubit2]
                pre_node_gate_type2 = node_int_all[pre_nodes[qubit2]][0]
            else:
                edge_index.append([0, node_index + 1])
            pre_nodes[qubit2] = node_index + 1
            node1 = get_graph_node(node1[0] + 1, node1[1], node1[2] - 1, pre_distance=pre_distance_1, pre_node=pre_node_gate_type1 + 1)
            node_int1 = [node1[0] + 1, node1[1], node1[2] - 1]
            node2 = get_graph_node(node2[0] + 1, node2[1], node2[2] - 1, pre_distance=pre_distance_2, pre_node=pre_node_gate_type2 + 1)
            node_int2 = [node2[0] + 1, node2[1], node2[2] - 1]
            nodes.append(node1)
            nodes.append(node2)
            temp_node_int.append(node_int1)
            temp_node_int.append(node_int2)
            depth_note[qubit1] = depth_max
            depth_note[qubit2] = depth_max
            edge_index.append([node_index, node_index + 1])
            edge_index.append([node_index + 1, node_index])
            node_index += 2
            i += 3
        for node in nodes:
            x.append(node)
        for node_int in temp_node_int:
            node_int_all.append(node_int)
    x = torch.tensor(x, dtype=torch.float)
    if right_rate >= 0.5:
        '''y = [1 for _ in range(node_count + 1)]'''
        y = 1
    else:
        '''y = [0 for _ in range(node_count + 1)]'''
        y = 0
    y = torch.tensor(y, dtype=torch.long)
    edge_index = torch.tensor(edge_index, dtype=torch.long)
    return Data(x=x, y=y, edge_index=edge_index.t().contiguous())


def trans_circuit_to_graph_delta(circuit_code, delta):  # 将电路转换为图数据结构
    codes = list(map(int, circuit_code))
    node_count = node_get(circuit_code)
    i = 0
    x = []
    '''y = 1 if right_rate >= 0.5 else 0'''

    '''if right_rate >= 0.5:
        y[1] = 1
    else:
        y[0] = 1'''
    node_int_all = []
    edge_index = []
    depth_note = []
    pre_nodes = []
    node_index = 1
    while i < len(codes):
        if i == 0:
            for j in range(codes[0]):
                depth_note.append(0)
                pre_nodes.append(-1)
            i += 1
            x.append([0 for _ in range(38)])
            node_int_all.append([])
            continue
        nodes = []
        temp_node_int = []
        gate_qubit_number = codes[i]
        i += 1

        if gate_qubit_number == 1:
            gate_type = codes[i] + 1
            if gate_type == 7:
                gate_type = 5
            qubit = codes[i+1]
            depth_note[qubit] += 1
            pre_node = pre_nodes[qubit]
            pre_distance = 0
            pre_node_gate_type = -1
            if pre_node != -1:
                edge_index.append([pre_node, node_index])
                pre_distance = 1
                pre_node_gate_type = node_int_all[pre_node][0]
            else:
                edge_index.append([0, node_index])
            node = get_graph_node(gate_type, qubit, depth_note[qubit] - 1, pre_distance=pre_distance, pre_node=pre_node_gate_type + 1)
            node_int = [gate_type, qubit, depth_note[qubit] - 1]
            temp_node_int.append(node_int)
            nodes.append(node)
            pre_nodes[qubit] = node_index
            node_index += 1
            i += 2
        elif gate_qubit_number == 2:
            gate_type = codes[i]
            qubit1 = codes[i+1]
            qubit2 = codes[i+2]
            node1 = [-1, qubit1]
            node2 = [0 if gate_type == 0 else -1, qubit2]
            depth_max = max(depth_note[qubit1], depth_note[qubit2]) + 1
            node1.append(depth_max)
            node2.append(depth_max)
            pre_distance_1 = 0
            pre_distance_2 = 0
            pre_node_gate_type1 = -1
            pre_node_gate_type2 = -1
            if pre_nodes[qubit1] != -1:
                edge_index.append([pre_nodes[qubit1], node_index])
                pre_distance_1 = depth_max - depth_note[qubit1]
                pre_node_gate_type1 = node_int_all[pre_nodes[qubit1]][0]
            else:
                edge_index.append([0, node_index])
            pre_nodes[qubit1] = node_index
            if pre_nodes[qubit2] != -1:
                edge_index.append([pre_nodes[qubit2], node_index + 1])
                pre_distance_2 = depth_max - depth_note[qubit2]
                pre_node_gate_type2 = node_int_all[pre_nodes[qubit2]][0]
            else:
                edge_index.append([0, node_index + 1])
            pre_nodes[qubit2] = node_index + 1
            node1 = get_graph_node(node1[0] + 1, node1[1], node1[2] - 1, pre_distance=pre_distance_1, pre_node=pre_node_gate_type1 + 1)
            node_int1 = [node1[0] + 1, node1[1], node1[2] - 1]
            node2 = get_graph_node(node2[0] + 1, node2[1], node2[2] - 1, pre_distance=pre_distance_2, pre_node=pre_node_gate_type2 + 1)
            node_int2 = [node2[0] + 1, node2[1], node2[2] - 1]
            nodes.append(node1)
            nodes.append(node2)
            temp_node_int.append(node_int1)
            temp_node_int.append(node_int2)
            depth_note[qubit1] = depth_max
            depth_note[qubit2] = depth_max
            edge_index.append([node_index, node_index + 1])
            edge_index.append([node_index + 1, node_index])
            node_index += 2
            i += 3
        for node in nodes:
            x.append(node)
        for node_int in temp_node_int:
            node_int_all.append(node_int)
    x = torch.tensor(x, dtype=torch.float)
    y = delta
    y = torch.tensor(y, dtype=torch.float)
    edge_index = torch.tensor(edge_index, dtype=torch.long)
    return Data(x=x, y=y, edge_index=edge_index.t().contiguous())


def circuit_cut(circuit_code):
    codes = list(map(int, circuit_code))
    cut_line = []
    i = 0
    while i < len(codes):
        recent_line = []
        if i == 0:
            recent_line = codes[0:1]
            i = 1
        else:
            recent_line = codes[i:i + codes[i] + 2]
            i = i + codes[i] + 2
        cut_line.append(recent_line)
    return cut_line


def wash_circuit(circuit_code):
    cut_line = circuit_cut(circuit_code)
    i = len(cut_line) - 1
    pre_width = cut_line[0][0]
    involve_multi = [0 for j in range(pre_width)]
    gate_number = 0
    multi_gate_number = 0
    measure_count = 0
    while i >= 1:
        recent_gate = list(map(int, cut_line[i]))
        if recent_gate[0] == 1 and recent_gate[1] == 6:
            measure_qubit = recent_gate[2]
            involve_multi[measure_qubit] = 1
            gate_number += 1
            i = i - 1
        else:
            bit_number = recent_gate[0]
            qubit_indexes = []
            for j in range(bit_number):  # 读取后几位的量子位位置
                qubit_indexes.append(recent_gate[j + 2])
            efficient = 0
            for qubit_index in qubit_indexes:
                if involve_multi[qubit_index] == 1:
                    gate_number += 1    # 修正门数
                    efficient = 1
                    if recent_gate[0] != 1:
                        multi_gate_number += 1  # 修正多位门数
                        for qubit_confirm in qubit_indexes:
                            involve_multi[qubit_confirm] = 1
                        break
            i = i - 1
            if efficient == 0:
                cut_line.pop(i + 1)

    # 检查空道
    qubits_involve = []
    remove_gate = []
    for gate in cut_line:
        if len(gate) == 1:
            qubits_involve = [0 for _ in range(gate[0])]
            continue
        elif gate[1] == 6:
            recent_m_qubit = gate[2]
            if qubits_involve[recent_m_qubit] == 0:
                remove_gate.append(gate)
            else:
                measure_count += 1
        else:
            bit = gate[0]
            if gate[1] != 6:
                for i in range(bit):
                    qubit = gate[2 + i]
                    qubits_involve[qubit] = 1
    for gate in remove_gate:
        cut_line.remove(gate)
        gate_number -= 1
    if len(cut_line) == 1:
        return '', 0, 0, 0, 0, 0

    bit_count = 0
    arrow = 0   # 偏移量
    for i, j, seq in zip(involve_multi, qubits_involve, range(len(involve_multi))):
        if i == 1 and j == 1:
            bit_count += 1
        else:
            arrow += 1
        if arrow != 0:
            for gate in cut_line:
                if len(gate) == 1:
                    continue
                qubit_number = gate[0]
                for q in range(qubit_number):
                    if gate[len(gate) - 1 - q] == seq:
                        gate[len(gate) - 1 - q] -= arrow
    cut_line[0][0] = bit_count  # 修正电路宽度


    depths = [0 for i in range(pre_width)]
    for gate in cut_line:
        if len(gate) == 1:
            continue
        max_depth = 0
        i = 0
        while i < len(gate):
            if i == 0 or i == 1:
                i = i + 1
                continue
            recent_qubit = gate[i]
            depths[recent_qubit] += 1
            if depths[recent_qubit] < max_depth:
                depths[recent_qubit] = max_depth
            else:
                max_depth = depths[recent_qubit]
            i = i + 1
    depth = max(depths)  # 修正电路深度
    string_code = ''
    for gate in cut_line:
        for code in gate:
            string_code += str(code)
    return string_code, cut_line[0][0], depth, gate_number, multi_gate_number, measure_count


def simulate_correct_rate(circuit_code, is_noise, all_p):
    c = read_circuit_code(circuit_code)
    sim = cirq.Simulator()
    result = sim.simulate(c)

    data_result = sim.run(c)

    random_channel = RandomNoise(all_p)
    c = read_circuit_code(circuit_code, is_noise, random_channel)
    random_param = [random_channel.random1, random_channel.random2, random_channel.random3]
    '''result_noise = sim.simulate(c)'''

    '''result = result.final_state_vector
    result_noise = result_noise.final_state_vector'''
    count = 0
    for i in range(1000):
        data_result_noise = sim.run(c)
        if data_result_noise == data_result:
            count += 1
    '''print(f'Real right rate:{count/1000}')
    correct_rate = 0
    for i in range(len(result)):
        if abs(result[i]) > 0:
            correct_rate = abs(result_noise[i])
            break'''
    return count/1000, random_param


def multi_circuit_get(width, depth, count_max):
    count = 0
    data = [['circuit_code', 'width', 'depth', 'gate_number', 'multi_gate_number', 'measurement_gate_number']]
    while count < count_max:
        string_code, width2, depth2, gate_number2, multi_gate_number2, measure_count2 = random_circuit_generate_no_gate_number(width, depth)
        row_data = [string_code, width2, depth2, gate_number2, multi_gate_number2, measure_count2]
        data.append(row_data)
        count += 1
        print()

        sys.stdout.write(f'\rCount:{count} circuit_code:{string_code} width:{width2} depth:{depth2} gate_number:{gate_number2}'
                         f' multi_gate_number:{multi_gate_number2} measure_gate_number:{measure_count2}')
        sys.stdout.flush()

    print()
    write_to_excel(f'satisfy_circuit{width}{depth}.xlsx', 'sheet1', 'info', data)


def main_circuit_transform():
    circuit_code = '4110120121120132111100113132100160161162163'
    c = read_circuit_code(circuit_code)
    print(c)
    result = cirq.Simulator().run(c, repetitions=100)
    print(result)

def main_circuit_transform_noise():
    circuit_code = '81271441221151271531471171051251361551513024013610020031301432062113130156213510715720352041143104160161162163164165166167'
    c = read_circuit_code(circuit_code, 1, RandomNoise())
    print(c)
    result = cirq.Simulator().simulate(c)
    print(result)


def main_circuit_cut():
    print(circuit_cut('320022102101131130162'))


def main_wash_circuit():
    c, width, depth, gate_number, multi_gate_number, measure_count = wash_circuit('4100122160161162163')
    print(f'{c} {width} {depth} {gate_number} {multi_gate_number} {measure_count}')


def main_right_rate():
    correct_rate, random_params = simulate_correct_rate('310011120122101160161162', 1, 0.1)
    print(f'correct rate:{correct_rate}')
    print(f'random params:{random_params}')


def main_trans_to_graph():
    data = trans_circuit_to_graph('321212121110120120160161162', right_rate=0.56)
    print(data)
    print(read_circuit_code('321212121110120120160161162'))


def main_node_get():
    count = node_get('321212121110120120160161162')
    print(count)


def dataset_generate(p_all):
    # 随机生成电路，输入satisfy_circuit.xlsx表格
    #multi_circuit_get(depth, width, gate_number, multi_gate_number, count)
    circuits_info = read_xls('unrepeated_satisfy_circuit.xlsx')
    circuits_info.pop(0)
    r = p_all
    random_channel = RandomNoise(r)
    sub_dataset_generate(circuits_info, random_channel)


def dataset_generate_by_certain_noise(random_x, random_y, random_z, output_filename):
    # 随机生成电路，输入satisfy_circuit.xlsx表格
    #multi_circuit_get(depth, width, gate_number, multi_gate_number, count)
    circuits_info = read_xls('D:\\workspace\\QuantumBase\\RandomCircuit\\TrainData\\origin_2w_five_gate_circuit.xlsx')
    circuits_info.pop(0)
    random_channel = RandomNoise()
    random_channel.random1 = random_x
    random_channel.random2 = random_y
    random_channel.random3 = random_z
    sub_dataset_generate_multi_sample(circuits_info, random_channel, output_filename)


def sub_dataset_generate(circuits_info, random_channel, output_filename):
    data = [
        ['circuit_code', 'width', 'depth', 'gate_number', 'multi_gate_number', 'measurement_gate_number', 'right_rate',
         'Pauli_error_x', 'Pauli_error_y', 'Pauli_error_z']]
    # 读取每一电路数据，获取电路，每个电路插入100次不同的噪声分布，总错误率小于0.5，获取对应的正确率。
    for circuit_info, all_count in zip(circuits_info, range(len(circuits_info))):
        circuit_code = circuit_info[0]
        c = read_circuit_code(circuit_code)
        sim = cirq.Simulator()
        result = sim.run(c)
        noise_c = c.with_noise(random_channel)
        right_count = 0
        for i in range(1000):
            noise_result = sim.run(noise_c)
            if result == noise_result:
                right_count += 1
            sys.stdout.write(f'\rC_Count:{all_count + 1} step_count:{i + 1} circuit_code:{circuit_code} right_count:{right_count} right_rate:{right_count / (i + 1)}')
            sys.stdout.flush()
        print()
        row_data = circuit_info
        row_data.append(right_count / 1000)
        row_data.append(random_channel.random1)
        row_data.append(random_channel.random2)
        row_data.append(random_channel.random3)
        data.append(row_data)
        if all_count % 100 == 0 and all_count != 0:
            write_to_excel(f'{output_filename}_{all_count}.xlsx', 'sheet1', 'info', data)
    write_to_excel(f'{output_filename}_ALL.xlsx', 'sheet1', 'info', data)


def sub_dataset_generate_multi_sample(circuits_info, random_channel, output_filename):
    data = [
        ['circuit_code', 'width', 'depth', 'gate_number', 'multi_gate_number', 'measurement_gate_number', 'right_rate',
         'Pauli_error_x', 'Pauli_error_y', 'Pauli_error_z', 'delta']]
    # 读取每一电路数据，获取电路，每个电路插入100次不同的噪声分布，总错误率小于0.5，获取对应的正确率。
    for circuit_info, all_count in zip(circuits_info, range(len(circuits_info))):
        circuit_code = circuit_info[0]
        c = read_circuit_code(circuit_code)
        sim = cirq.Simulator()
        result = sim.run(c)
        noise_c = c.with_noise(random_channel)
        right_count = 0
        false_count = []
        result_sequence = []
        for i in range(1000):
            noise_result = sim.run(noise_c)
            if result == noise_result:
                right_count += 1
                continue
            else:
                if len(false_count) == 0:
                    false_count.append(1)
                    result_sequence.append(noise_result)
                else:
                    judge = 0
                    for index, result in zip(range(len(false_count)), result_sequence):
                        if noise_result == result:
                            judge = 1
                            false_count[index] += 1
                            break
                    if judge == 0:
                        false_count.append(1)
                        result_sequence.append(noise_result)
            second_count = max(false_count)
            error_delta = right_count / (i + 1) - second_count / (i + 1)
            sys.stdout.write(f'\rC_Count:{all_count + 1} step_count:{i + 1} circuit_code:{circuit_code} right_count:{right_count} second_count:{second_count} right_rate:{right_count / (i + 1)} error_delta:{error_delta}')
            sys.stdout.flush()
        sys.stdout.write(
            f'\rC_Count:{all_count + 1} step_count:{1000} circuit_code:{circuit_code} right_count:{right_count} second_count:{max(false_count)} right_rate:{right_count / (i + 1)} error_delta:{(right_count - max(false_count)) / 1000}')
        sys.stdout.flush()
        print()
        row_data = circuit_info
        row_data.append(right_count / 1000)
        row_data.append(random_channel.random1)
        row_data.append(random_channel.random2)
        row_data.append(random_channel.random3)
        row_data.append((right_count - max(false_count)) / 1000)
        data.append(row_data)
        if all_count % 100 == 0 and all_count != 0:
            write_to_excel(f'{output_filename}_{all_count}.xlsx', 'sheet1', 'info', data)
    write_to_excel(f'{output_filename}_ALL.xlsx', 'sheet1', 'info', data)


def read_xls(location):
    excel = load_workbook(location)
    # 获取sheet：
    table = excel.get_sheet_by_name('sheet1')  # 通过表名获取
    # 获取行数和列数：
    rows = table.max_row  # 获取行数
    cols = table.max_column  # 获取列数
    # 获取单元格值：
    data = []
    for i in range(rows):
        row_data = []
        for j in range(cols):
            row_data.append(table.cell(row=i + 1, column=j + 1).value)  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
        data.append(row_data)
    return data


def get_binary_list(number, bit):
    if number >= bit:
        raise Exception('Number bigger than bit or equals bit.')
    if number < 0:
        raise Exception('Number less than 0.')
    binary_list = [0 for _ in range(bit)]
    binary_list[number] = 1
    return binary_list


def get_graph_node(gate_type, qubit, depth, pre_distance, pre_node):
    gate_type_list = get_binary_list(gate_type, 6)
    qubit_list = get_binary_list(qubit, 8)
    depth_list = get_binary_list(depth, 9)
    pre_distance_list = get_binary_list(pre_distance, 9)
    pre_node_list = get_binary_list(pre_node, 6)
    node = gate_type_list + qubit_list + depth_list + pre_distance_list + pre_node_list
    return node


if __name__ == '__main__':
    #main_circuit_transform()
    #main_trans_to_graph()
    #main_node_get()
    #multi_circuit_get(5, 5, 2000)
    #multi_circuit_get(9, 9, 15000)
    #dataset_generate(0.1)
    #main_wash_circuit()
    dataset_generate_by_certain_noise(0.029428343, 0.057264254, 0.013307403, '0.1_e1_delta_2w')
